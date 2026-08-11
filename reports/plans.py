import io
import re
from datetime import date, datetime

import openpyxl
import pandas as pd

from reports._utils import autofit_columns, parse_date, write_headers

STATUS_ORDER = [
    "Awarded - Active",
    "Awarded - Closed",
    "Application Submitted",
    "LOI Submitted",
    "Application In Progress",
    "LOI In Progress",
    "Planned",
    "Researching",
    "Declined",
    "Abandoned",
]

COLUMN_MAP = {
    "Year": "Year",
    "Funder name": "Funder",
    "Project": "Fund",
    "Request Purpose": "Purpose",
    "Status": "Status",
    "Amount requested": "Request",
    "Amount awarded": "Award",
    "Expected notification date": "Notif Expected",
    "Notification date": "Notif Received",
    "Next task deadline": "Next Task/Deadline",
}

OUTPUT_COLUMNS = [
    "Year",
    "Funder",
    "Fund",
    "Purpose",
    "Status",
    "Request",
    "Award",
    "Notif Expected",
    "Notif Received",
    "Next Task/Deadline",
]

# Excel number formats applied in _build_worksheet.
_CURRENCY_FORMAT = '"$"#,##0'
_DATE_FORMAT = "m/d/yy"


def _current_fiscal_year(today: date) -> int:
    """Return the current fiscal year number (July 1-June 30 convention)."""
    return today.year + 1 if today.month >= 7 else today.year


def _tab_sort_key(year: int, current_fy: int) -> tuple[int, int]:
    """Order fiscal-year tabs: current first, then future ascending, then past
    descending — e.g. with a current FY of 2027: 2027, 2028, 2029, 2026, 2025."""
    return (0, year) if year >= current_fy else (1, -year)


def _parse_year(val):
    """Parse a Year value to an int.

    Tolerates fiscal-year labels like "FY 2026" / "FY2026" (used by clients
    whose fiscal years cross the calendar year) by disregarding the "FY"
    prefix and reading the numeric year. Returns None if no year is found.
    """
    if pd.isna(val):
        return None
    s = re.sub(r"^\s*FY\s*", "", str(val).strip(), flags=re.IGNORECASE)
    try:
        return int(float(s))
    except ValueError:
        return None


def _parse_amount(val):
    if pd.isna(val) or not str(val).strip():
        return None
    s = re.sub(r"[^\d.]", "", str(val))
    try:
        return float(s) if s else None
    except ValueError:
        return None


def _build_worksheet(ws, df):
    write_headers(ws, OUTPUT_COLUMNS)

    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, col_name in enumerate(OUTPUT_COLUMNS, 1):
            val = row.get(col_name)
            cell = ws.cell(row=row_idx, column=col_idx)

            if col_name in ("Notif Expected", "Notif Received"):
                if isinstance(val, datetime):
                    cell.value = val
                    cell.number_format = _DATE_FORMAT
                else:
                    cell.value = None
            elif col_name in ("Request", "Award"):
                cell.value = val
                if val is not None:
                    cell.number_format = _CURRENCY_FORMAT
            else:
                cell.value = None if (isinstance(val, float) and pd.isna(val)) else val

    autofit_columns(ws, OUTPUT_COLUMNS)


def _load_data(csv_bytes: bytes) -> list[tuple[str, dict[int, pd.DataFrame]]]:
    """Parse and validate a Plans CSV.

    Returns a list of (client, fy_map) tuples, one per unique client, where
    fy_map maps fiscal-year numbers to DataFrames with OUTPUT_COLUMNS columns.
    Rows are sorted by STATUS_ORDER rank then Fund name. Every fiscal year in
    the file is included; fy_map is ordered current fiscal year first, then
    future years ascending, then past years descending (e.g. 2027, 2028, 2029,
    2026, 2025).
    """
    try:
        df = pd.read_csv(io.BytesIO(csv_bytes))
    except (pd.errors.ParserError, UnicodeDecodeError, pd.errors.EmptyDataError) as e:
        raise ValueError(f"This file couldn't be read as a CSV: {e}")

    if df.empty:
        raise ValueError("The uploaded file has no data rows.")

    missing = [c for c in list(COLUMN_MAP.keys()) + ["Client"] if c not in df.columns]
    if missing:
        raise ValueError(f"CSV is missing expected columns: {', '.join(missing)}")

    clients = df["Client"].fillna("Unknown")
    out = df[list(COLUMN_MAP.keys())].rename(columns=COLUMN_MAP).copy()
    out["_client"] = clients.values

    out["Request"] = out["Request"].apply(_parse_amount)
    out["Award"] = out["Award"].apply(_parse_amount)
    out["Notif Expected"] = out["Notif Expected"].apply(parse_date)
    out["Notif Received"] = out["Notif Received"].apply(parse_date)

    out["Year"] = out["Year"].apply(_parse_year)
    bad_year_count = int(out["Year"].isna().sum())
    if bad_year_count:
        raise ValueError(
            f"{bad_year_count} row(s) have non-numeric Year values — please check the CSV."
        )
    out["Year"] = out["Year"].astype(int)

    status_rank = {s: i for i, s in enumerate(STATUS_ORDER)}
    out["_status_rank"] = out["Status"].map(status_rank).fillna(len(STATUS_ORDER))
    out = out.sort_values(["_status_rank", "Fund"], na_position="last")
    current_fy = _current_fiscal_year(date.today())

    results = []
    for client, client_df in out.groupby("_client"):
        years = sorted(client_df["Year"].unique(), key=lambda y: _tab_sort_key(y, current_fy))
        fy_map = {
            year: client_df[client_df["Year"] == year][OUTPUT_COLUMNS].reset_index(drop=True)
            for year in years
        }
        results.append((client, fy_map))
    return results


def generate(csv_bytes: bytes) -> list[tuple[str, bytes]]:
    results = []
    for client, fy_map in _load_data(csv_bytes):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for year, year_df in fy_map.items():
            ws = wb.create_sheet(title=f"FY {year}")
            _build_worksheet(ws, year_df)

        wb.active = wb[f"FY {next(iter(fy_map))}"]

        xlsx_buf = io.BytesIO()
        wb.save(xlsx_buf)
        results.append((client, xlsx_buf.getvalue()))

    return results
