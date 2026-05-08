from datetime import datetime

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")

_DATE_FORMATS = ("%b %d, %Y", "%m/%d/%Y", "%Y-%m-%d")


def parse_date(val):
    if pd.isna(val) or not str(val).strip():
        return None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(str(val).strip(), fmt)
        except ValueError:
            continue
    return None


def write_headers(ws, columns):
    for col_idx, header in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def autofit_columns(ws, columns):
    for col_idx, header in enumerate(columns, 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(header)
        for row_idx in range(2, ws.max_row + 1):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                if isinstance(cell_val, datetime):
                    max_len = max(max_len, 10)
                else:
                    max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
