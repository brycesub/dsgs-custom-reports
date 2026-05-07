import io
import re
import zipfile
from datetime import date, datetime

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

STATUS_ORDER = [
    "Awarded - Active",
    "Awarded - Closed",
    "Application Submitted",
    "LOI Submitted",
    "Application In Progress",
    "LOI In Progress",
    "Planned",
    "Researching",
    "Declined",
    "Abandoned",
]

COLUMN_MAP = {
    "Year": "Year",
    "Funder name": "Funder",
    "Project": "Fund",
    "Request Purpose": "Purpose",
    "Status": "Status",
    "Amount requested": "Request",
    "Amount awarded": "Award",
    "Expected notification date": "Notif Expected",
    "Notification date": "Notif Received",
    "Next task deadline": "Next Task/Deadline",
}

OUTPUT_COLUMNS = [
    "Year", "Funder", "Fund", "Purpose", "Status",
    "Request", "Award", "Notif Expected", "Notif Received", "Next Task/Deadline",
]

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _parse_amount(val):
    if pd.isna(val) or not str(val).strip():
        return None
    s = re.sub(r"[^\d.]", "", str(val))
    try:
        return float(s) if s else None
    except ValueError:
        return None


def _parse_date(val):
    if pd.isna(val) or not str(val).strip():
        return None
    for fmt in ("%b %d, %Y", "%m/%d/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(val).strip(), fmt)
        except ValueError:
            continue
    return None


def _extract_task_date(val):
    if not val or pd.isna(val):
        return None
    m = re.search(r"(\d{2}/\d{2}/\d{4})", str(val))
    if m:
        try:
            return datetime.strptime(m.group(1), "%m/%d/%Y")
        except ValueError:
            return None
    return None


def _build_worksheet(ws, df):
    for col_idx, header in enumerate(OUTPUT_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"

    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, col_name in enumerate(OUTPUT_COLUMNS, 1):
            val = row.get(col_name)
            cell = ws.cell(row=row_idx, column=col_idx)

            if col_name in ("Notif Expected", "Notif Received"):
                if isinstance(val, datetime):
                    cell.value = val
                    cell.number_format = "MM/DD/YYYY"
                else:
                    cell.value = None
            elif col_name in ("Request", "Award"):
                cell.value = val if val is not None else None
                if val is not None:
                    cell.number_format = "#,##0"
            else:
                cell.value = None if (isinstance(val, float) and pd.isna(val)) else val

    for col_idx, header in enumerate(OUTPUT_COLUMNS, 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(header)
        for row_idx in range(2, ws.max_row + 1):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                # Don't use datetime string length for width — use formatted length
                if isinstance(cell_val, datetime):
                    max_len = max(max_len, 10)
                else:
                    max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)


def generate(csv_bytes: bytes) -> bytes:
    df = pd.read_csv(io.BytesIO(csv_bytes))

    missing = [c for c in list(COLUMN_MAP.keys()) + ["Client"] if c not in df.columns]
    if missing:
        raise ValueError(f"CSV is missing expected columns: {', '.join(missing)}")

    clients = df["Client"].fillna("Unknown")
    out = df[list(COLUMN_MAP.keys())].rename(columns=COLUMN_MAP).copy()
    out["_client"] = clients.values

    out["Request"] = out["Request"].apply(_parse_amount)
    out["Award"] = out["Award"].apply(_parse_amount)
    out["Notif Expected"] = out["Notif Expected"].apply(_parse_date)
    out["Notif Received"] = out["Notif Received"].apply(_parse_date)

    status_rank = {s: i for i, s in enumerate(STATUS_ORDER)}
    out["_status_rank"] = out["Status"].map(status_rank).fillna(len(STATUS_ORDER))
    out["_task_date"] = out["Next Task/Deadline"].apply(_extract_task_date)

    out = out.sort_values(["_status_rank", "_task_date"], na_position="last")
    out = out[out["Year"].astype(int) >= date.today().year]

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for client, client_df in out.groupby("_client"):
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            current_year = date.today().year
            future_label = f"{current_year + 1}+"

            cur_df = client_df[client_df["Year"].astype(int) == current_year]
            fut_df = client_df[client_df["Year"].astype(int) > current_year]

            ws_cur = wb.create_sheet(title=str(current_year))
            _build_worksheet(ws_cur, cur_df[OUTPUT_COLUMNS])

            if not fut_df.empty:
                ws_fut = wb.create_sheet(title=future_label)
                _build_worksheet(ws_fut, fut_df[OUTPUT_COLUMNS])

            wb.active = wb[str(current_year)]

            xlsx_buf = io.BytesIO()
            wb.save(xlsx_buf)
            xlsx_buf.seek(0)
            zf.writestr(f"{client}_report.xlsx", xlsx_buf.read())

    zip_buf.seek(0)
    return zip_buf.read()
