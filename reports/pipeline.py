import io
import re
from datetime import date, datetime

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

COLUMN_MAP = {
    "Year": "Year",
    "Project": "Project",
    "Funder name": "Funder",
    "Task type": "Task Type",
    "Task Title": "Task Name",
    "Due Date": "Due Date",
}

OUTPUT_COLUMNS = ["Year", "Project", "Funder", "Task Type", "Task Name", "Due Date"]

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _parse_date(val):
    if pd.isna(val) or not str(val).strip():
        return None
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%b %d, %Y"):
        try:
            return datetime.strptime(str(val).strip(), fmt)
        except ValueError:
            continue
    return None


_PIPELINE_FILENAME_RE = re.compile(r"^Pipeline \d{8} .+\.csv$", re.IGNORECASE)


def _extract_client(filename: str) -> str:
    if not _PIPELINE_FILENAME_RE.match(filename):
        raise ValueError(
            f"Filename must match 'Pipeline YYYYMMDD ClientName.csv' — got: {filename!r}"
        )
    stem = re.sub(r"\.csv$", "", filename, flags=re.IGNORECASE)
    parts = stem.split(" ", 2)
    return parts[2]


def generate(csv_bytes: bytes, filename: str) -> tuple[str, bytes]:
    try:
        df = pd.read_csv(io.BytesIO(csv_bytes))
    except pd.errors.ParserError:
        raise ValueError("This file couldn't be read as a CSV — please check the format.")

    if df.empty:
        raise ValueError("The uploaded file has no data rows.")

    missing = [c for c in COLUMN_MAP if c not in df.columns]
    if missing:
        raise ValueError(f"CSV is missing expected columns: {', '.join(missing)}")

    client = _extract_client(filename)
    sheet_name = f"Pipeline - {date.today().isoformat()}"

    out = df[list(COLUMN_MAP.keys())].rename(columns=COLUMN_MAP).copy()
    out["Due Date"] = out["Due Date"].apply(_parse_date)
    out = out.sort_values("Due Date", na_position="last")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title=sheet_name)

    for col_idx, header in enumerate(OUTPUT_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"

    for row_idx, (_, row) in enumerate(out.iterrows(), 2):
        for col_idx, col_name in enumerate(OUTPUT_COLUMNS, 1):
            val = row.get(col_name)
            cell = ws.cell(row=row_idx, column=col_idx)
            if col_name == "Due Date":
                if isinstance(val, datetime):
                    cell.value = val
                    cell.number_format = "MM/DD/YYYY"
                else:
                    cell.value = None
            else:
                cell.value = None if (isinstance(val, float) and pd.isna(val)) else val

    for col_idx, header in enumerate(OUTPUT_COLUMNS, 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(header)
        for row_idx in range(2, ws.max_row + 1):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                if isinstance(cell_val, datetime):
                    max_len = max(max_len, 10)
                else:
                    max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    xlsx_buf = io.BytesIO()
    wb.save(xlsx_buf)
    return client, xlsx_buf.getvalue()
