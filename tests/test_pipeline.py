import io
from datetime import date, datetime
from unittest.mock import MagicMock, patch

import openpyxl
import pytest

from reports.pipeline import _extract_client, generate

FIXED_TODAY = date(2026, 5, 8)
FIXED_TODAY_STR = "2026-05-08"

_HEADER = "Year,Project,Funder name,Task type,Task Title,Due Date"
_VALID_FILENAME = "Pipeline 20260508 HS.csv"


def _csv(*rows):
    return ("\n".join([_HEADER] + list(rows))).encode()


def _row(
    year=2026,
    project="Test Project",
    funder="Test Funder",
    task_type="Report",
    task_title="My Task",
    due_date="2026-05-01",
):
    return f"{year},{project},{funder},{task_type},{task_title},{due_date}"


@pytest.fixture(autouse=True)
def freeze_today():
    mock = MagicMock()
    mock.today.return_value = FIXED_TODAY
    with patch("reports.pipeline.date", mock):
        yield


# ---------------------------------------------------------------------------
# _extract_client
# ---------------------------------------------------------------------------


class TestExtractClient:
    def test_single_word_client(self):
        assert _extract_client("Pipeline 20260508 HS.csv") == "HS"

    def test_multi_word_client(self):
        assert _extract_client("Pipeline 20260508 My Client.csv") == "My Client"

    def test_raises_for_missing_date(self):
        with pytest.raises(ValueError, match="Filename must match"):
            _extract_client("Pipeline HS.csv")

    def test_raises_for_wrong_prefix(self):
        with pytest.raises(ValueError, match="Filename must match"):
            _extract_client("Plans 20260508 HS.csv")

    def test_raises_for_non_csv_extension(self):
        with pytest.raises(ValueError, match="Filename must match"):
            _extract_client("Pipeline 20260508 HS.xlsx")

    def test_csv_extension_case_insensitive(self):
        assert _extract_client("Pipeline 20260508 HS.CSV") == "HS"


# ---------------------------------------------------------------------------
# Happy path
# ---------------------------------------------------------------------------


class TestGenerateHappyPath:
    def test_returns_client_and_bytes(self):
        client, xlsx_bytes = generate(_csv(_row()), _VALID_FILENAME)
        assert client == "HS"
        assert isinstance(xlsx_bytes, bytes) and len(xlsx_bytes) > 0

    def test_sheet_name_includes_today(self):
        _, xlsx_bytes = generate(_csv(_row()), _VALID_FILENAME)
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        assert f"Pipeline - {FIXED_TODAY_STR}" in wb.sheetnames

    def test_headers_written_to_row_1(self):
        _, xlsx_bytes = generate(_csv(_row()), _VALID_FILENAME)
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        assert ws.cell(1, 1).value == "Year"
        assert ws.cell(1, 5).value == "Task Name"
        assert ws.cell(1, 6).value == "Due Date"

    def test_data_written_to_row_2(self):
        _, xlsx_bytes = generate(_csv(_row(project="HS/Edu", funder="Big Fund")), _VALID_FILENAME)
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        assert ws.cell(2, 2).value == "HS/Edu"  # Project
        assert ws.cell(2, 3).value == "Big Fund"  # Funder

    def test_due_date_cell_is_datetime(self):
        _, xlsx_bytes = generate(_csv(_row(due_date="2026-05-01")), _VALID_FILENAME)
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        assert isinstance(ws.cell(2, 6).value, datetime)

    def test_unparseable_due_date_is_none(self):
        _, xlsx_bytes = generate(_csv(_row(due_date="not a date")), _VALID_FILENAME)
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        assert ws.cell(2, 6).value is None


class TestExtraColumnsIgnored:
    def test_collaborator_attachments_status_not_in_output(self):
        header = (
            "Year,Project,Funder name,Task type,Task Title,Due Date,"
            "Collaborator(s),Attachments,Status"
        )
        rows = [
            (
                "FY 2026,HS/Edu,Scripps Family Fund,Report,Fin Report,07/31/2026,"
                "Diana Silver,file.pdf,Incomplete"
            ),
            "FY 2027,HS/Edu,Big Fund,General,LOI Draft,08/15/2026,Diana Silver,,Not Started",
        ]
        csv_bytes = ("\n".join([header] + rows)).encode()
        _, xlsx_bytes = generate(csv_bytes, _VALID_FILENAME)
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        assert ws.max_column == 6
        assert [ws.cell(1, c).value for c in range(1, 7)] == [
            "Year",
            "Project",
            "Funder",
            "Task Type",
            "Task Name",
            "Due Date",
        ]
        values = [
            ws.cell(r, c).value
            for r in range(1, ws.max_row + 1)
            for c in range(1, ws.max_column + 1)
        ]
        for sentinel in ("Diana Silver", "file.pdf", "Incomplete", "Not Started"):
            assert sentinel not in values


# ---------------------------------------------------------------------------
# Sort order
# ---------------------------------------------------------------------------


class TestSortOrder:
    def test_rows_sorted_by_due_date_ascending(self):
        csv_bytes = _csv(
            _row(task_title="Later Task", due_date="2026-06-01"),
            _row(task_title="Earlier Task", due_date="2026-04-01"),
        )
        _, xlsx_bytes = generate(csv_bytes, _VALID_FILENAME)
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        assert ws.cell(2, 5).value == "Earlier Task"  # Task Name
        assert ws.cell(3, 5).value == "Later Task"

    def test_null_due_dates_sorted_last(self):
        csv_bytes = _csv(
            _row(task_title="No Date", due_date=""),
            _row(task_title="Has Date", due_date="2026-04-01"),
        )
        _, xlsx_bytes = generate(csv_bytes, _VALID_FILENAME)
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        assert ws.cell(2, 5).value == "Has Date"
        assert ws.cell(3, 5).value == "No Date"


# ---------------------------------------------------------------------------
# Validation errors
# ---------------------------------------------------------------------------


class TestValidationErrors:
    def test_raises_for_bad_filename(self):
        with pytest.raises(ValueError, match="Filename must match"):
            generate(_csv(_row()), "badfilename.csv")

    def test_raises_for_missing_column(self):
        csv_bytes = b"Year,Project,Funder name\n2026,P,F"
        with pytest.raises(ValueError, match="missing expected columns"):
            generate(csv_bytes, _VALID_FILENAME)

    def test_raises_for_header_only_csv(self):
        with pytest.raises(ValueError, match="no data rows"):
            generate((_HEADER + "\n").encode(), _VALID_FILENAME)
