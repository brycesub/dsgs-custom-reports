from datetime import datetime

import openpyxl
import pandas as pd

from reports._utils import autofit_columns, parse_date, write_headers


class TestParseDate:
    def test_parses_month_day_year_format(self):
        assert parse_date("Jan 01, 2026") == datetime(2026, 1, 1)

    def test_parses_slash_format(self):
        assert parse_date("01/15/2026") == datetime(2026, 1, 15)

    def test_parses_iso_format(self):
        assert parse_date("2026-03-20") == datetime(2026, 3, 20)

    def test_returns_none_for_empty_string(self):
        assert parse_date("") is None

    def test_returns_none_for_float_nan(self):
        assert parse_date(float("nan")) is None

    def test_returns_none_for_pd_na(self):
        assert parse_date(pd.NA) is None

    def test_returns_none_for_unrecognized_format(self):
        assert parse_date("not a date") is None

    def test_strips_whitespace(self):
        assert parse_date("  2026-03-20  ") == datetime(2026, 3, 20)


class TestWriteHeaders:
    def setup_method(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active

    def test_writes_header_values(self):
        write_headers(self.ws, ["Alpha", "Beta", "Gamma"])
        assert self.ws.cell(1, 1).value == "Alpha"
        assert self.ws.cell(1, 2).value == "Beta"
        assert self.ws.cell(1, 3).value == "Gamma"

    def test_bold_font(self):
        write_headers(self.ws, ["Col"])
        assert self.ws.cell(1, 1).font.bold is True

    def test_white_font_color(self):
        write_headers(self.ws, ["Col"])
        assert "FFFFFF" in self.ws.cell(1, 1).font.color.rgb

    def test_blue_solid_fill(self):
        write_headers(self.ws, ["Col"])
        fill = self.ws.cell(1, 1).fill
        assert fill.fill_type == "solid"
        assert "4472C4" in fill.fgColor.rgb

    def test_center_alignment(self):
        write_headers(self.ws, ["Col"])
        assert self.ws.cell(1, 1).alignment.horizontal == "center"

    def test_freeze_panes_at_a2(self):
        write_headers(self.ws, ["Col"])
        assert self.ws.freeze_panes == "A2"


class TestAutofitColumns:
    def setup_method(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active

    def test_width_at_least_header_plus_two(self):
        header = "ShortHdr"
        self.ws.cell(1, 1, value=header)
        autofit_columns(self.ws, [header])
        assert self.ws.column_dimensions["A"].width >= len(header) + 2

    def test_width_expands_for_long_cell_value(self):
        header = "Col"
        long_val = "A very long cell value here"
        self.ws.cell(1, 1, value=header)
        self.ws.cell(2, 1, value=long_val)
        autofit_columns(self.ws, [header])
        assert self.ws.column_dimensions["A"].width >= len(long_val) + 2

    def test_width_capped_at_50(self):
        header = "Col"
        self.ws.cell(1, 1, value=header)
        self.ws.cell(2, 1, value="x" * 100)
        autofit_columns(self.ws, [header])
        assert self.ws.column_dimensions["A"].width == 50

    def test_datetime_value_counted_as_10_chars(self):
        header = "Date"
        self.ws.cell(1, 1, value=header)
        self.ws.cell(2, 1, value=datetime(2026, 1, 1))
        autofit_columns(self.ws, [header])
        assert self.ws.column_dimensions["A"].width >= 10 + 2
