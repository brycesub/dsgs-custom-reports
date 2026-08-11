import io
from datetime import date
from unittest.mock import MagicMock, patch

import openpyxl
import pytest

from reports.plans import generate

# All tests freeze today to 2026-01-15 so year-filter logic is deterministic.
FIXED_TODAY = date(2026, 1, 15)
CURRENT_YEAR = 2026
FUTURE_YEAR = 2027

_HEADER = (
    "Year,Client,Funder name,Project,Request Purpose,Status,"
    "Amount requested,Amount awarded,Expected notification date,"
    "Notification date,Next task deadline"
)


def _csv(*rows):
    return ("\n".join([_HEADER] + list(rows))).encode()


def _row(
    year=CURRENT_YEAR,
    client="TestClient",
    funder="Test Funder",
    project="Test Project",
    purpose="General",
    status="Planned",
    request="10000",
    award="",
    notif_expected="",
    notif_received="",
    next_deadline="",
):
    return (
        f"{year},{client},{funder},{project},{purpose},{status},"
        f"{request},{award},{notif_expected},{notif_received},{next_deadline}"
    )


@pytest.fixture(autouse=True)
def freeze_today():
    mock = MagicMock()
    mock.today.return_value = FIXED_TODAY
    with patch("reports.plans.date", mock):
        yield


# ---------------------------------------------------------------------------
# Happy path
# ---------------------------------------------------------------------------


class TestGenerateHappyPath:
    def test_returns_one_result_per_client(self):
        results = generate(_csv(_row(client="A"), _row(client="B")))
        assert {r[0] for r in results} == {"A", "B"}

    def test_result_is_non_empty_bytes(self):
        _, xlsx_bytes = generate(_csv(_row()))[0]
        assert isinstance(xlsx_bytes, bytes) and len(xlsx_bytes) > 0

    def test_client_name_in_result(self):
        client, _ = generate(_csv(_row(client="Acme Corp")))[0]
        assert client == "Acme Corp"

    def test_current_year_sheet_always_created(self):
        _, xlsx_bytes = generate(_csv(_row(year=CURRENT_YEAR)))[0]
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        assert f"FY {CURRENT_YEAR}" in wb.sheetnames

    def test_future_sheet_created_when_future_rows_exist(self):
        _, xlsx_bytes = generate(_csv(_row(year=CURRENT_YEAR), _row(year=FUTURE_YEAR)))[0]
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        assert f"FY {FUTURE_YEAR}" in wb.sheetnames

    def test_no_future_sheet_when_no_future_rows(self):
        _, xlsx_bytes = generate(_csv(_row(year=CURRENT_YEAR)))[0]
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        assert len(wb.sheetnames) == 1

    def test_current_year_sheet_is_active(self):
        _, xlsx_bytes = generate(_csv(_row(year=CURRENT_YEAR), _row(year=FUTURE_YEAR)))[0]
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        assert wb.active.title == f"FY {CURRENT_YEAR}"

    def test_headers_written_to_row_1(self):
        _, xlsx_bytes = generate(_csv(_row()))[0]
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb[f"FY {CURRENT_YEAR}"]
        # OUTPUT_COLUMNS: Year, Funder, Fund, Purpose, Status, Request, Award, ...
        assert ws.cell(1, 1).value == "Year"
        assert ws.cell(1, 2).value == "Funder"
        assert ws.cell(1, 3).value == "Fund"
        assert ws.cell(1, 5).value == "Status"

    def test_data_written_to_row_2(self):
        _, xlsx_bytes = generate(_csv(_row(funder="Big Foundation", project="My Fund")))[0]
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb[f"FY {CURRENT_YEAR}"]
        assert ws.cell(2, 2).value == "Big Foundation"  # Funder
        assert ws.cell(2, 3).value == "My Fund"  # Fund


# ---------------------------------------------------------------------------
# Year inclusion
# ---------------------------------------------------------------------------


class TestYearInclusion:
    def test_past_year_rows_get_their_own_sheet(self):
        # Two rows: one past (2025), one current (2026)
        _, xlsx_bytes = generate(_csv(_row(year=2025), _row(year=CURRENT_YEAR)))[0]
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        assert wb.sheetnames == [f"FY {CURRENT_YEAR}", "FY 2025"]
        assert wb[f"FY {CURRENT_YEAR}"].max_row == 2  # header + 1 row
        assert wb["FY 2025"].max_row == 2

    def test_all_past_rows_still_produce_a_workbook(self):
        _, xlsx_bytes = generate(_csv(_row(year=2024), _row(year=2025)))[0]
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        # Newest past year first; with no current/future year it is also active.
        assert wb.sheetnames == ["FY 2025", "FY 2024"]
        assert wb.active.title == "FY 2025"

    def test_future_rows_go_to_future_sheet(self):
        _, xlsx_bytes = generate(_csv(_row(year=FUTURE_YEAR)))[0]
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        assert f"FY {FUTURE_YEAR}" in wb.sheetnames
        assert wb[f"FY {FUTURE_YEAR}"].max_row == 2  # header + 1 row


# ---------------------------------------------------------------------------
# Sort order
# ---------------------------------------------------------------------------


class TestSortOrder:
    def test_status_rank_applied(self):
        # "Awarded - Active" ranks 0; "Planned" ranks 6 — should appear first
        csv_bytes = _csv(
            _row(status="Planned", project="ZZZ Fund"),
            _row(status="Awarded - Active", project="AAA Fund"),
        )
        _, xlsx_bytes = generate(csv_bytes)[0]
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb[f"FY {CURRENT_YEAR}"]
        assert ws.cell(2, 5).value == "Awarded - Active"
        assert ws.cell(3, 5).value == "Planned"

    def test_alphabetical_fund_within_same_status(self):
        csv_bytes = _csv(
            _row(status="Planned", project="Zoo Fund"),
            _row(status="Planned", project="Alpha Fund"),
        )
        _, xlsx_bytes = generate(csv_bytes)[0]
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb[f"FY {CURRENT_YEAR}"]
        assert ws.cell(2, 3).value == "Alpha Fund"  # Fund col
        assert ws.cell(3, 3).value == "Zoo Fund"


# ---------------------------------------------------------------------------
# Amount parsing
# ---------------------------------------------------------------------------


class TestAmountParsing:
    def test_dollar_sign_and_commas_stripped(self):
        _, xlsx_bytes = generate(_csv(_row(request="$10500", award="$3000")))[0]
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb[f"FY {CURRENT_YEAR}"]
        assert ws.cell(2, 6).value == pytest.approx(10500.0)  # Request
        assert ws.cell(2, 7).value == pytest.approx(3000.0)  # Award

    def test_empty_amount_written_as_none(self):
        _, xlsx_bytes = generate(_csv(_row(request="", award="")))[0]
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb[f"FY {CURRENT_YEAR}"]
        assert ws.cell(2, 6).value is None
        assert ws.cell(2, 7).value is None


# ---------------------------------------------------------------------------
# Number formats
# ---------------------------------------------------------------------------


class TestNumberFormats:
    def test_request_award_use_currency_format(self):
        _, xlsx_bytes = generate(_csv(_row(request="175000", award="3000")))[0]
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb[f"FY {CURRENT_YEAR}"]
        assert ws.cell(2, 6).number_format == '"$"#,##0'  # Request
        assert ws.cell(2, 7).number_format == '"$"#,##0'  # Award

    def test_notif_dates_use_short_date_format(self):
        _, xlsx_bytes = generate(
            _csv(_row(notif_expected="2026-06-30", notif_received="2026-04-21"))
        )[0]
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb[f"FY {CURRENT_YEAR}"]
        assert ws.cell(2, 8).number_format == "m/d/yy"  # Notif Expected
        assert ws.cell(2, 9).number_format == "m/d/yy"  # Notif Received


# ---------------------------------------------------------------------------
# Validation errors
# ---------------------------------------------------------------------------


class TestValidationErrors:
    def test_raises_for_missing_column(self):
        csv_bytes = b"Year,Funder name,Project\n2026,Some Funder,Some Project"
        with pytest.raises(ValueError, match="missing expected columns"):
            generate(csv_bytes)

    def test_raises_for_header_only_csv(self):
        with pytest.raises(ValueError, match="no data rows"):
            generate(_HEADER.encode() + b"\n")

    def test_raises_for_non_numeric_year(self):
        with pytest.raises(ValueError, match="non-numeric Year"):
            generate(_csv(_row(year="TBD")))


# ---------------------------------------------------------------------------
# Fiscal-year ("FY 2026") handling
# ---------------------------------------------------------------------------


class TestFiscalYearParsing:
    """Some clients (KCRep, Hartford Stage, WEN) use 'FY 2026' for fiscal years
    that cross the calendar year. The 'FY' prefix should be disregarded and the
    row treated as its numeric year."""

    def test_fy_prefixed_current_year_included(self):
        _, xlsx_bytes = generate(_csv(_row(year="FY 2026")))[0]
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb[f"FY {CURRENT_YEAR}"]
        assert ws.max_row == 2  # header + 1 row
        assert ws.cell(2, 1).value == CURRENT_YEAR

    def test_fy_prefixed_future_year_goes_to_future_sheet(self):
        _, xlsx_bytes = generate(_csv(_row(year="FY 2027")))[0]
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        assert wb[f"FY {FUTURE_YEAR}"].max_row == 2

    def test_fy_prefixed_past_year_kept_on_its_own_sheet(self):
        _, xlsx_bytes = generate(_csv(_row(year="FY 2025"), _row(year="FY 2026")))[0]
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        assert wb.sheetnames == [f"FY {CURRENT_YEAR}", "FY 2025"]
        assert wb["FY 2025"].cell(2, 1).value == 2025

    def test_mixed_fy_and_plain_years_sort_together(self):
        # A plain 2026 and an "FY 2026" should land on the same sheet.
        csv_bytes = _csv(
            _row(year="FY 2026", status="Planned", project="B Fund"),
            _row(year=2026, status="Planned", project="A Fund"),
        )
        _, xlsx_bytes = generate(csv_bytes)[0]
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb[f"FY {CURRENT_YEAR}"]
        assert ws.max_row == 3  # header + 2 rows
        assert ws.cell(2, 3).value == "A Fund"
        assert ws.cell(3, 3).value == "B Fund"

    def test_fy_no_space(self):
        _, xlsx_bytes = generate(_csv(_row(year="FY2026")))[0]
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb[f"FY {CURRENT_YEAR}"]
        assert ws.cell(2, 1).value == CURRENT_YEAR


class TestFiscalYearTabs:
    def test_one_sheet_per_fiscal_year_ascending(self):
        csv_bytes = _csv(
            _row(year=2027, project="B Fund", status="Planned"),
            _row(year=2026, project="A Fund", status="Planned"),
            _row(year=2029, project="C Fund", status="Planned"),
        )
        _, xlsx_bytes = generate(csv_bytes)[0]
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        assert wb.sheetnames == ["FY 2026", "FY 2027", "FY 2029"]

    def test_august_current_fy_first_and_active(self):
        with patch("reports.plans.date") as mock_date:
            mock_date.today.return_value = date(2026, 8, 6)
            _, xlsx_bytes = generate(
                _csv(
                    _row(year=2026, project="Past", status="Planned"),
                    _row(year=2027, project="Current", status="Planned"),
                    _row(year=2028, project="Future", status="Planned"),
                )
            )[0]
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        assert wb.sheetnames == ["FY 2027", "FY 2028", "FY 2026"]
        assert wb.active.title == "FY 2027"
        assert wb["FY 2027"].max_row == 2  # header + 1

    def test_future_years_ascend_then_past_years_descend(self):
        with patch("reports.plans.date") as mock_date:
            mock_date.today.return_value = date(2026, 8, 6)  # current FY 2027
            _, xlsx_bytes = generate(
                _csv(
                    _row(year=2024, status="Planned"),
                    _row(year=2029, status="Planned"),
                    _row(year=2026, status="Planned"),
                    _row(year=2028, status="Planned"),
                    _row(year=2027, status="Planned"),
                    _row(year=2025, status="Planned"),
                )
            )[0]
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        assert wb.sheetnames == [
            "FY 2027",
            "FY 2028",
            "FY 2029",
            "FY 2026",
            "FY 2025",
            "FY 2024",
        ]

    def test_sort_order_preserved_within_a_past_year_sheet(self):
        csv_bytes = _csv(
            _row(year=2024, status="Planned", project="ZZZ Fund"),
            _row(year=2024, status="Awarded - Active", project="AAA Fund"),
        )
        _, xlsx_bytes = generate(csv_bytes)[0]
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["FY 2024"]
        assert ws.cell(2, 5).value == "Awarded - Active"
        assert ws.cell(3, 5).value == "Planned"
